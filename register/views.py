from django.shortcuts import render
from .models import *

from datetime import datetime

from django.views.generic import TemplateView
from django.shortcuts import render
from django.shortcuts import redirect
from django.http import HttpResponse

# libreria para el manejo de archivos xls
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, Fill
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
from smtplib import SMTPRecipientsRefused
from django.core.mail import EmailMultiAlternatives
from django.template.loader import get_template
from django.template import Context
from register.models import Delivery


def index(request):
  template = 'index.html'
  context = {
  'title': "GUARDIAN",
  }
  return render(request, template, context)


def mail(request):
  template = 'mail.html'
  deliveries = Delivery.objects.all()

  context = {
    'trailer': deliveries[0].trailer,
    'cons': deliveries,
  }
  return render(request, template, context)


def sendmail(request):

  deliveries = Delivery.objects.all()

  new_context = {
  'trailer': deliveries[0].trailer,
  'cons': deliveries,
  }
  template = get_template('mail.html')
  html_content = template.render(new_context)


  fromaddr = "itzli2000@gmail.com"
  toaddr = "itzli2000@msn.com"
  # Error on create html template
  msg = EmailMultiAlternatives(subject="Información de ventas 'ICE'",
                               body=html_content,)
  msg.attach(MIMEText('Reporte_General_De_Entregas.xlsx', attachment.read(), 'text/xlsx'))
  msg['From'] = fromaddr
  msg['To'] = toaddr
  msg['Subject'] = "Información de ventas 'ICE'"
  body = html_content
  msg.attach(MIMEText(body, 'html'))
  # msg.attach('invoice.csv', csvfile.getvalue(), 'text/csv')
  server = smtplib.SMTP('smtp.gmail.com', 587)
  server.starttls()
  server.login(fromaddr, "molinona&9")
  text = msg.as_string()
  server.sendmail(fromaddr, toaddr, text)
  server.quit()

  return render(request, template, context)


class ReportDelivery(TemplateView):
  """
  Vista basada en clases que retorna un archivo de reporte de formato xls (Excel)
      HttpResponse de content_type='application/ms-excel'
  Para obtener archivo, debe ser llamado desde petición get
  """

  def get(self, request, *args, **kwargs):
    """
    Funcionamiento:
        Para la generación de los campos, unicamente se deben cargar los modelos
        y sus campos respectivos, así como sus inner joins que se requieren para
        mostrarse en el excel.
    """
    # Se cargan los modelos para insertar la información
    deliverys = Delivery.objects.all()
    # Se el objeto xls
    workbook = Workbook()
    # Contador para iniciar desde la fila 4 a colocar la información de los modelos
    count = 3
    # Activa la hoja de excel 1 para trabajar
    ws1 = workbook.active
    # Inserta contenido a celdas
    ws1.title = "Reporte de entregas"
    ws1['A1'] = 'Reporte general de entregas'
    # ws1.column_dimensions['D3'].width = 150
    # Aplica estilos a las celdas
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A1:G1')

    # Se definen los encabezados de las columnas
    ws1['A2'] = 'ID'
    ws1['B2'] = 'Numero'
    ws1['C2'] = 'Trailer'
    ws1['D2'] = 'Fecha'
    ws1['E2'] = 'Unidades'
    ws1['F2'] = 'Producto'
    ws1['G2'] = 'Total'
    ws1.column_dimensions["D"].width = 25
    c = ws1['D2']
    c.font = Font(size=12)
    for delivery in deliverys:
      # Itera celdas y añade contenido
      ws1.cell(row=count, column=1, value=delivery.id)
      ws1.cell(row=count, column=2, value=delivery.number)
      ws1.cell(row=count, column=3, value=delivery.trailer)
      ws1.cell(row=count, column=4, value=delivery.date)
      ws1.cell(row=count, column=5, value=delivery.units)
      ws1.cell(row=count, column=6, value=delivery.product)
      ws1.cell(row=count, column=7, value=delivery.total)

    file_name = 'Reporte_General_De_Usuarios_{0}.xlsx'.format(datetime.now().strftime("%I-%M%p_%d-%m-%Y"))
    response = HttpResponse(content_type='application/ms-excel')
    content = 'attachment; filename={0}'.format(file_name)
    response['Content-Disposition'] = content
    workbook.save(response)
    return response
